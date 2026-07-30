"""
Crea el nuevo Excel "ATM Catering Deportivo.xlsx" con las 3 pestañas:
  - CALENDARIO TEMPORADA: partidos + contactos via fórmula
  - EQUIPOS VISITANTES: base de datos maestra de contactos
  - REGISTRO: seguimiento semanal

Uso: python sports_catering/crear_excel_catering.py
Importa contactos del Excel actual si existe.
"""
import sys
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── Colores corporativos ──────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1A1A2E"   # azul oscuro ATM
COLOR_HEADER_FG   = "FFFFFF"   # blanco
COLOR_SUBHEAD_BG  = "2D2D4E"   # azul medio
COLOR_FORMULA_BG  = "EBF5FB"   # azul muy claro (celdas con fórmula)
COLOR_MANUAL_BG   = "FFF3CD"   # amarillo claro (relleno manual)
COLOR_GREEN_BG    = "D5F5E3"   # verde claro (con contacto)
COLOR_RED_BG      = "FADBD8"   # rojo claro (sin contacto)
COLOR_ALT_ROW     = "F8F9FA"   # gris muy claro (fila par)
COLOR_BORDER      = "CCCCCC"

# ── Utilidades de estilo ──────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def _border_thin() -> Border:
    s = Side(style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def _apply_header(cell, text: str, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FG, size=10):
    cell.value = text
    cell.font = _font(bold=True, color=fg, size=size)
    cell.fill = _fill(bg)
    cell.border = _border_thin()
    cell.alignment = _center()

def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width

def _set_row_height(ws, row: int, height: float):
    ws.row_dimensions[row].height = height


# ── Hoja 1: EQUIPOS VISITANTES ────────────────────────────────────────────────
def build_equipos_visitantes(ws, existing_contacts: list[dict]):
    """
    Tabla maestra de contactos. Cada fila = un equipo visitante.
    Columnas: A=RIVAL, B=CONTACTO, C=CARGO, D=TELÉFONO, E=EMAIL, F=NOTAS
    """
    ws.sheet_properties.tabColor = "2E86C1"
    _set_row_height(ws, 1, 32)

    headers = [
        ("RIVAL / EQUIPO VISITANTE", 40),
        ("PERSONA DE CONTACTO", 22),
        ("CARGO", 18),
        ("TELÉFONO", 16),
        ("EMAIL", 32),
        ("NOTAS", 28),
    ]
    for col, (title, width) in enumerate(headers, start=1):
        _apply_header(ws.cell(1, col), title)
        _set_col_width(ws, col, width)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F1"

    # Importar contactos existentes
    for i, c in enumerate(existing_contacts, start=2):
        ws.cell(i, 1).value = c.get("nombre", "")
        ws.cell(i, 2).value = c.get("contacto", "")
        ws.cell(i, 3).value = c.get("cargo", "")
        ws.cell(i, 4).value = c.get("telefono", "")
        ws.cell(i, 5).value = c.get("email", "")
        ws.cell(i, 6).value = c.get("notas", "")

        # Filas alternas
        fill = _fill(COLOR_ALT_ROW) if i % 2 == 0 else PatternFill()
        for col in range(1, 7):
            cell = ws.cell(i, col)
            cell.border = _border_thin()
            cell.alignment = _left()
            cell.font = _font()
            if i % 2 == 0:
                cell.fill = _fill(COLOR_ALT_ROW)

    # Fila de instrucción al final
    last = len(existing_contacts) + 3
    ws.cell(last, 1).value = "← Añade aquí nuevos equipos visitantes"
    ws.cell(last, 1).font = _font(italic=True, color="888888")

    # Nombre de tabla para referencias
    ws.title = "EQUIPOS VISITANTES"


# ── Hoja 2: CALENDARIO TEMPORADA ─────────────────────────────────────────────
def build_calendario(ws, ev_sheet_name: str = "EQUIPOS VISITANTES"):
    """
    Calendario completo de la temporada. Columnas:
    A=DÍA SEMANA (fórmula), B=FECHA, C=HORA, D=DEPORTE, E=EQUIPO LOCAL,
    F=CIUDAD, G=RIVAL, H=PABELLÓN, I=CONTACTO (fórmula), J=TELÉFONO (fórmula),
    K=EMAIL (fórmula), L=CONTACTADO, M=NOTAS

    Las columnas I, J, K se rellenan solas con BUSCARV desde EQUIPOS VISITANTES.
    """
    ws.title = "CALENDARIO TEMPORADA"
    ws.sheet_properties.tabColor = "1E8449"
    _set_row_height(ws, 1, 14)
    _set_row_height(ws, 2, 32)

    # Fila 1: cabecera de grupo (3 secciones)
    ws.merge_cells("A1:H1")
    ws.merge_cells("I1:K1")
    ws.merge_cells("L1:M1")
    _apply_header(ws.cell(1, 1), "DATOS DEL PARTIDO", bg=COLOR_HEADER_BG)
    _apply_header(ws.cell(1, 9), "CONTACTO (automático desde EQUIPOS VISITANTES)", bg=COLOR_SUBHEAD_BG)
    _apply_header(ws.cell(1, 12), "SEGUIMIENTO", bg="555555")

    # Fila 2: cabeceras de columna
    col_headers = [
        ("DÍA",         8),
        ("FECHA",       13),
        ("HORA",        8),
        ("DEPORTE",     14),
        ("EQUIPO LOCAL",20),
        ("CIUDAD",      10),
        ("RIVAL",       36),
        ("PABELLÓN",    28),
        ("PERSONA",     22),   # fórmula
        ("TELÉFONO",    16),   # fórmula
        ("EMAIL",       32),   # fórmula
        ("CONTACTADO",  14),
        ("NOTAS",       24),
    ]
    for col, (title, width) in enumerate(col_headers, start=1):
        is_formula = col in (9, 10, 11)
        bg = COLOR_SUBHEAD_BG if is_formula else COLOR_HEADER_BG
        _apply_header(ws.cell(2, col), title, bg=bg)
        _set_col_width(ws, col, width)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:M2"

    # Pre-rellenar 400 filas con fórmulas (toda la temporada cabe holgado)
    for row in range(3, 403):
        alt = row % 2 == 0

        # Col A: día de la semana (fórmula desde FECHA en B)
        cell_a = ws.cell(row, 1)
        cell_a.value = f'=IFERROR(IF(B{row}="","",TEXT(B{row},"DDDD")),"")'
        cell_a.fill = _fill(COLOR_FORMULA_BG)
        cell_a.font = _font(italic=True, color="555555")
        cell_a.border = _border_thin()
        cell_a.alignment = _center()

        # Cols B-H: datos a rellenar manualmente / por script
        manual_cols = {
            2: ("fecha",    "DD/MM/YYYY", _center()),
            3: ("hora",     "HH:MM",      _center()),
            4: ("deporte",  None,         _center()),
            5: ("equipo",   None,         _left()),
            6: ("ciudad",   None,         _center()),
            7: ("rival",    None,         _left()),
            8: ("pabellon", None,         _left()),
        }
        for col, (_, fmt, align) in manual_cols.items():
            cell = ws.cell(row, col)
            cell.border = _border_thin()
            cell.alignment = align
            cell.font = _font()
            if alt:
                cell.fill = _fill(COLOR_ALT_ROW)
            if fmt == "DD/MM/YYYY":
                cell.number_format = "DD/MM/YYYY"
            elif fmt == "HH:MM":
                cell.number_format = "HH:MM"

        # Col I: PERSONA DE CONTACTO (BUSCARV desde EQUIPOS VISITANTES col B)
        cell_i = ws.cell(row, 9)
        cell_i.value = (
            f'=IFERROR(IF(G{row}="","",VLOOKUP(G{row},\'EQUIPOS VISITANTES\'!$A:$F,2,0)),"")'
        )
        cell_i.fill = _fill(COLOR_FORMULA_BG)
        cell_i.font = _font(italic=True, color="1A5276")
        cell_i.border = _border_thin()
        cell_i.alignment = _left()

        # Col J: TELÉFONO (BUSCARV col D)
        cell_j = ws.cell(row, 10)
        cell_j.value = (
            f'=IFERROR(IF(G{row}="","",VLOOKUP(G{row},\'EQUIPOS VISITANTES\'!$A:$F,4,0)),"")'
        )
        cell_j.fill = _fill(COLOR_FORMULA_BG)
        cell_j.font = _font(italic=True, color="1A5276")
        cell_j.border = _border_thin()
        cell_j.alignment = _center()

        # Col K: EMAIL (BUSCARV col E)
        cell_k = ws.cell(row, 11)
        cell_k.value = (
            f'=IFERROR(IF(G{row}="","",VLOOKUP(G{row},\'EQUIPOS VISITANTES\'!$A:$F,5,0)),"")'
        )
        cell_k.fill = _fill(COLOR_FORMULA_BG)
        cell_k.font = _font(italic=True, color="1A5276")
        cell_k.border = _border_thin()
        cell_k.alignment = _left()

        # Col L: CONTACTADO (desplegable: —/PDTE./SI/NO)
        cell_l = ws.cell(row, 12)
        cell_l.value = f'=IF(G{row}="","","PDTE.")'
        cell_l.border = _border_thin()
        cell_l.alignment = _center()
        cell_l.font = _font()
        if alt:
            cell_l.fill = _fill(COLOR_ALT_ROW)

        # Col M: NOTAS
        cell_m = ws.cell(row, 13)
        cell_m.border = _border_thin()
        cell_m.alignment = _left()
        cell_m.font = _font()
        if alt:
            cell_m.fill = _fill(COLOR_ALT_ROW)

    # Leyenda debajo
    leg_row = 404
    ws.cell(leg_row, 1).value = "ℹ️  Las columnas azules (PERSONA, TELÉFONO, EMAIL) se rellenan automáticamente cuando el equipo está en la pestaña EQUIPOS VISITANTES."
    ws.cell(leg_row, 1).font = _font(italic=True, color="555555", size=9)
    ws.merge_cells(f"A{leg_row}:M{leg_row}")


# ── Hoja 3: REGISTRO ──────────────────────────────────────────────────────────
def build_registro(ws):
    """
    Registro de contactos semanales. Columnas:
    A=N°, B=FECHA CONTACTO, C=DEPORTE, D=EQUIPO LOCAL, E=CIUDAD,
    F=RIVAL, G=FECHA PARTIDO, H=EMAIL ENVIADO, I=RESPUESTA, J=PEDIDO, K=IMPORTE, L=NOTAS
    """
    ws.title = "REGISTRO"
    ws.sheet_properties.tabColor = "E74C3C"
    _set_row_height(ws, 1, 32)

    headers = [
        ("N°",             6),
        ("FECHA CONTACTO", 16),
        ("DEPORTE",        14),
        ("EQUIPO LOCAL",   20),
        ("CIUDAD",         10),
        ("RIVAL",          36),
        ("FECHA PARTIDO",  16),
        ("EMAIL ENVIADO",  14),
        ("RESPUESTA",      14),
        ("PEDIDO",         12),
        ("IMPORTE €",      12),
        ("NOTAS",          30),
    ]
    for col, (title, width) in enumerate(headers, start=1):
        _apply_header(ws.cell(1, col), title)
        _set_col_width(ws, col, width)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:L1"

    # 300 filas pre-rellenadas con fórmula de N° correlativo
    for row in range(2, 302):
        ws.cell(row, 1).value = f"=IF(F{row}=\"\",\"\",ROW()-1)"
        ws.cell(row, 1).font = _font(color="555555", italic=True)
        ws.cell(row, 1).fill = _fill(COLOR_FORMULA_BG)
        ws.cell(row, 1).border = _border_thin()
        ws.cell(row, 1).alignment = _center()

        for col in range(2, 13):
            cell = ws.cell(row, col)
            cell.border = _border_thin()
            cell.alignment = _left() if col in (3,4,5,6,12) else _center()
            cell.font = _font()
            if row % 2 == 0:
                cell.fill = _fill(COLOR_ALT_ROW)
            if col in (2, 7):
                cell.number_format = "DD/MM/YYYY"
            if col == 11:
                cell.number_format = '#,##0.00 €'


# ── Builder principal ─────────────────────────────────────────────────────────
def create_excel(output_path: str, existing_contacts: list[dict] = None):
    wb = openpyxl.Workbook()

    # Eliminar hoja por defecto
    del wb[wb.sheetnames[0]]

    # Crear hojas en orden lógico de uso
    ws_cal = wb.create_sheet("CALENDARIO TEMPORADA")
    ws_ev  = wb.create_sheet("EQUIPOS VISITANTES")
    ws_reg = wb.create_sheet("REGISTRO")

    build_equipos_visitantes(ws_ev, existing_contacts or [])
    build_calendario(ws_cal)
    build_registro(ws_reg)

    wb.save(output_path)
    print(f"Excel creado en: {output_path}")
    return output_path


# ── Importar contactos del Excel actual ───────────────────────────────────────
def load_contacts_from_existing(existing_excel_path: str) -> list[dict]:
    """Lee la pestaña EQUIPOS VISITANTES del Excel actual."""
    try:
        wb = openpyxl.load_workbook(existing_excel_path, data_only=True)
        ws = wb["EQUIPOS VISITANTES"]
        contacts = []
        for row in range(2, ws.max_row + 1):
            nombre = ws.cell(row, 1).value
            if not nombre:
                continue
            email1 = ws.cell(row, 6).value
            email2 = ws.cell(row, 7).value
            contacts.append({
                "nombre":   str(nombre).strip(),
                "contacto": str(ws.cell(row, 3).value or "").strip(),
                "cargo":    str(ws.cell(row, 4).value or "").strip(),
                "telefono": str(ws.cell(row, 5).value or "").strip() if ws.cell(row, 5).value else "",
                "email":    str(email1 or email2 or "").strip(),
                "notas":    "",
            })
        return contacts
    except Exception as e:
        print(f"  Aviso: no se pudo importar contactos existentes ({e}). Se creará vacío.")
        return []


if __name__ == "__main__":
    import shutil, os

    # Intentar leer contactos del Excel actual
    current_excel = r"C:\Users\marco\ATM burgers\ATM-Documentación - Documentos\EVENTOS\REGISTRO EVENTOS.xlsx"
    temp_copy = r"C:\Users\marco\AppData\Local\Temp\registro_read.xlsx"

    contacts = []
    try:
        shutil.copy2(current_excel, temp_copy)
        contacts = load_contacts_from_existing(temp_copy)
        print(f"Importados {len(contacts)} contactos del Excel actual.")
    except Exception as e:
        print(f"No se pudo copiar el Excel actual: {e}")
        print("El nuevo Excel se creará con EQUIPOS VISITANTES vacío.")

    # Ruta de salida
    output = r"C:\Users\marco\ATM burgers\ATM-Documentación - Documentos\EVENTOS\ATM Catering Deportivo.xlsx"
    create_excel(output, contacts)
    print("Listo. Puedes abrir el archivo en Excel.")
