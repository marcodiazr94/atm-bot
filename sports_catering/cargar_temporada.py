"""
ATM Catering Deportivo — Carga inicial de temporada.

Descarga TODOS los partidos en casa de la temporada para cada equipo asturiano
y los escribe en la pestaña CALENDARIO TEMPORADA del Excel "ATM Catering Deportivo.xlsx".

Uso (una vez al año, en agosto/septiembre):
    python -m sports_catering.cargar_temporada [--dry-run] [--force]

--dry-run : muestra los partidos encontrados sin escribir en el Excel
--force   : sobreescribe el calendario aunque ya haya datos (útil para refrescar)
"""
import argparse
import shutil
import sys
import time
import unicodedata
from datetime import datetime, timezone

import openpyxl
from openpyxl.cell.cell import MergedCell

from sports_catering.config import load_config
from sports_catering.scrapers import espn, feb_competiciones, feb_envivo, rfef_juvenil
from sports_catering.utils.asturias_filter import init_from_config, is_asturian

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# Columnas de datos en CALENDARIO TEMPORADA (1-indexed)
# Las fórmulas están en cols A(1), I(9), J(10), K(11), L(12) — no tocar
COL_FECHA   = 2   # B
COL_HORA    = 3   # C
COL_DEPORTE = 4   # D
COL_EQUIPO  = 5   # E
COL_CIUDAD  = 6   # F
COL_RIVAL   = 7   # G
COL_PABELLON= 8   # H
FIRST_DATA_ROW = 3


def fetch_all_season(team: dict) -> list[dict]:
    """Obtiene TODOS los partidos de la temporada del equipo (sin filtro de fechas)."""
    source = team.get("source", "")
    name = team["name"]

    if source == "espn":
        team_id = team.get("espn_id")
        league = team.get("espn_league", "")
        if not team_id:
            return []
        matches = espn.get_team_schedule(int(team_id), league)
        time.sleep(0.5)
        return matches

    elif source == "feb_competiciones":
        autonomy_id = team.get("feb_autonomy_id")
        comp_id = team.get("feb_team_id")
        if not autonomy_id or not comp_id:
            return []
        # Sin filtro de fecha — devuelve todos
        matches = feb_competiciones.get_calendar(autonomy_id, comp_id, name)
        time.sleep(1)
        return matches

    elif source == "feb_envivo":
        slug = team.get("feb_slug", "")
        group = team.get("feb_group", 2)
        year = team.get("feb_season_year", datetime.now().year)
        keyword = team.get("feb_team_keyword", name)
        if not slug:
            return []
        matches = feb_envivo.get_calendar(slug, group, year, keyword)
        time.sleep(1)
        return matches

    elif source == "rfef_juvenil":
        url = team.get("rfef_url", "")
        keyword = team.get("rfef_team_keyword", name)
        if not url:
            return []
        matches = rfef_juvenil.get_calendar(url, keyword)
        time.sleep(1)
        return matches

    return []


def _hora_float(dt: datetime) -> float:
    return dt.hour + dt.minute / 60.0


def _norm(s: str) -> str:
    return unicodedata.normalize("NFD", s.upper()).encode("ascii", "ignore").decode()


def _is_home(match: dict, team: dict) -> bool:
    """Comprueba que el equipo asturiano juega de local."""
    home = _norm(match.get("home_team", ""))
    candidates = [
        team["name"],
        team.get("feb_team_keyword", ""),
        team.get("rfef_team_keyword", ""),
    ]
    for candidate in candidates:
        if not candidate:
            continue
        n = _norm(candidate)
        if n in home or home in n:
            return True
    return False


def _find_last_data_row(ws) -> int:
    """Última fila con datos reales en la hoja CALENDARIO (col G = RIVAL)."""
    last = FIRST_DATA_ROW - 1
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        if ws.cell(row, COL_RIVAL).value is not None:
            last = row
    return last


def _load_existing_rivals(ws) -> set[tuple]:
    """Devuelve set de (rival_upper, fecha_date) ya en el calendario."""
    existing = set()
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        rival = ws.cell(row, COL_RIVAL).value
        fecha = ws.cell(row, COL_FECHA).value
        if rival and fecha:
            fecha_d = fecha.date() if isinstance(fecha, datetime) else fecha
            existing.add((str(rival).upper().strip(), fecha_d))
    return existing


def _clear_calendar(ws):
    """Borra todos los datos del calendario (sin tocar filas de cabecera ni fórmulas)."""
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        for col in [COL_FECHA, COL_HORA, COL_DEPORTE, COL_EQUIPO, COL_CIUDAD, COL_RIVAL, COL_PABELLON]:
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _write_match(ws, row: int, match: dict):
    """Escribe un partido en la fila indicada. No toca las columnas de fórmula."""
    fecha = match["date"].replace(tzinfo=None)
    ws.cell(row, COL_FECHA).value = fecha
    ws.cell(row, COL_FECHA).number_format = "DD/MM/YYYY"

    hora = _hora_float(match["date"])
    ws.cell(row, COL_HORA).value = match["date"].replace(
        hour=int(hora), minute=round((hora % 1) * 60), second=0, microsecond=0
    ).time()
    ws.cell(row, COL_HORA).number_format = "HH:MM"

    ws.cell(row, COL_DEPORTE).value  = match.get("deporte", "")
    ws.cell(row, COL_EQUIPO).value   = match.get("equipo", "")
    ws.cell(row, COL_CIUDAD).value   = match.get("ciudad", "")
    ws.cell(row, COL_RIVAL).value    = match.get("rival", "")
    ws.cell(row, COL_PABELLON).value = match.get("venue", "")


def run(args):
    config = load_config()
    teams = config["teams"]

    catering_excel = config.get(
        "catering_excel_path",
        r"C:\Users\marco\ATM burgers\ATM-Documentación - Documentos\EVENTOS\ATM Catering Deportivo.xlsx",
    )

    now = datetime.now()
    print(f"\n{'='*65}")
    print(f"  ATM CATERING — CARGA DE TEMPORADA {now.year}/{now.year+1}")
    print(f"  {now.strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*65}\n")

    init_from_config(teams)

    # ── 1. Recoger partidos de todas las fuentes ───────────────────────────────
    all_home_matches: list[dict] = []
    manual_teams: list[dict] = []

    for team in teams:
        source = team.get("source", "")
        if source == "manual":
            manual_teams.append(team)
            continue

        print(f"Descargando: {team['name']} ({team['sport']}, {source})...")
        try:
            raw = fetch_all_season(team)
        except Exception as e:
            print(f"  ERROR: {e}")
            raw = []

        # Filtrar partidos en casa con visitante no asturiano
        added = 0
        for m in raw:
            if not isinstance(m.get("date"), datetime):
                continue
            if not _is_home(m, team):
                continue
            away = m.get("away_team", "")
            if is_asturian(away):
                continue
            all_home_matches.append({
                **m,
                "equipo":  team["name"],
                "ciudad":  team["city"],
                "deporte": team["sport"],
                "venue":   team.get("venue", ""),
                "rival":   away,
            })
            added += 1

        print(f"  {len(raw)} partidos → {added} en casa vs visitantes externos\n")

    # Ordenar por fecha
    all_home_matches.sort(key=lambda m: m["date"].replace(tzinfo=None))

    print(f"{'─'*65}")
    print(f"  TOTAL: {len(all_home_matches)} partidos en casa encontrados")
    print(f"{'─'*65}\n")

    if args.dry_run:
        print("[ DRY RUN — no se modifica el Excel ]\n")
        for m in all_home_matches:
            print(f"  {m['date'].strftime('%d/%m/%Y %H:%M')}  {m['equipo']:<25} vs {m['rival']}")
        _print_manual(manual_teams)
        return

    if not all_home_matches:
        print("Sin partidos encontrados. Verifica que la temporada ya ha comenzado.")
        _print_manual(manual_teams)
        return

    # ── 2. Abrir el Excel de catering ─────────────────────────────────────────
    print(f"Abriendo Excel: {catering_excel}")
    temp = catering_excel.replace(".xlsx", "_tmp.xlsx")
    try:
        shutil.copy2(catering_excel, temp)
    except FileNotFoundError:
        print(f"  ERROR: no se encuentra el archivo.\n  Ejecuta primero crear_excel_catering.py")
        sys.exit(1)

    wb = openpyxl.load_workbook(temp, data_only=False)
    ws = wb["CALENDARIO TEMPORADA"]

    # ── 3. Gestión de duplicados ───────────────────────────────────────────────
    if args.force:
        print("  --force: borrando datos existentes del calendario...")
        _clear_calendar(ws)
        existing = set()
    else:
        existing = _load_existing_rivals(ws)
        if existing:
            print(f"  {len(existing)} partidos ya en el calendario — se omitirán duplicados.")

    # ── 4. Escribir partidos nuevos ────────────────────────────────────────────
    last_row = _find_last_data_row(ws)
    next_row = last_row + 1
    written = 0
    skipped = 0

    for m in all_home_matches:
        fecha_d = m["date"].replace(tzinfo=None).date()
        key = (m["rival"].upper().strip(), fecha_d)
        if key in existing:
            skipped += 1
            continue

        if next_row > 402:   # límite de filas pre-formateadas
            print("  AVISO: Se alcanzó el límite de 400 partidos. Aumenta el Excel.")
            break

        _write_match(ws, next_row, m)
        existing.add(key)
        next_row += 1
        written += 1

    # ── 5. Guardar — salvar en temporal, luego reemplazar el original ─────────
    bak = catering_excel.replace(".xlsx", ".bak")
    try:
        shutil.copy2(catering_excel, bak)
    except Exception:
        pass

    wb.save(temp)
    try:
        shutil.move(temp, catering_excel)
    except PermissionError:
        print(f"\n  ERROR: El archivo está abierto en Excel. Ciérralo y vuelve a intentarlo.")
        print(f"  Los datos se han guardado temporalmente en: {temp}")
        sys.exit(1)

    print(f"\n  Escritos:  {written} partidos nuevos")
    if skipped:
        print(f"  Omitidos:  {skipped} duplicados")
    print(f"  Backup guardado: {bak}")
    print(f"\n  Excel actualizado: {catering_excel}\n")

    _print_manual(manual_teams)
    print("Proceso completado. Abre el Excel y completa los equipos manuales.\n")


def _print_manual(manual_teams: list[dict]):
    if not manual_teams:
        return
    print(f"{'─'*65}")
    print("  EQUIPOS A RELLENAR MANUALMENTE EN EL EXCEL:")
    print("  (busca su calendario en el link y añade los partidos a mano)")
    print()
    for t in manual_teams:
        url = t.get("manual_url", "sin URL configurada")
        print(f"  • {t['name']:<40} {t['sport']}")
        print(f"    {url}")
    print()


def main():
    parser = argparse.ArgumentParser(
        description="ATM Catering — Carga calendario completo de temporada"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Solo muestra los partidos encontrados, no modifica el Excel"
    )
    parser.add_argument(
        "--force", action="store_true",
        help="Borra el calendario existente y lo recarga desde cero"
    )
    args = parser.parse_args()
    run(args)


if __name__ == "__main__":
    main()
