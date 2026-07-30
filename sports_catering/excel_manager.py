"""
Gestión del Excel REGISTRO EVENTOS.xlsx:
- Leer partidos existentes y contactos conocidos
- Escribir nuevos partidos en la hoja REGISTRO (respetando fórmulas pre-existentes)
"""
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

# Columnas en la hoja REGISTRO (1-indexed)
COL_DIA = 1        # A: fórmula día semana (pre-rellenada)
COL_NUM = 2        # B: N° (pre-rellenada con número correlativo)
COL_PEDIDO = 3     # C: PEDIDO → "PDTE." para nuevos
COL_A = 4          # D: siempre vacío
COL_CIUDAD = 5     # E: LOCALIDAD
COL_DEPORTE = 6    # F: DEPORTE
COL_EQUIPO = 7     # G: EQUIPO (equipo local asturiano)
COL_RIVAL = 8      # H: RIVAL / ORGANIZADOR
COL_FECHA = 9      # I: FECHA (datetime)
COL_HORA = 10      # J: HORA (decimal, ej. 18.0 = 18:00)
COL_LUGAR = 11     # K: LUGAR (pabellón)
COL_TEL = 12       # L: Teléfono
COL_EMAIL = 13     # M: Email
# Cols N-T: fórmulas pre-rellenadas (año, mes, día, confirmación, etc.)


def load_workbook(path: str):
    """Carga el workbook con fórmulas (nunca data_only=True)."""
    return openpyxl.load_workbook(path, data_only=False)


def _find_last_data_row(ws) -> int:
    """Devuelve el índice de la última fila con datos reales (col H = RIVAL no vacío)."""
    last = 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, COL_RIVAL).value is not None:
            last = row
    return last


def load_existing_matches(wb) -> list[dict]:
    """Lee todos los partidos del REGISTRO con datos reales."""
    ws = wb["REGISTRO"]
    matches = []
    for row in range(2, ws.max_row + 1):
        rival = ws.cell(row, COL_RIVAL).value
        if not rival:
            continue
        fecha = ws.cell(row, COL_FECHA).value
        equipo = ws.cell(row, COL_EQUIPO).value
        matches.append({
            "row": row,
            "num": ws.cell(row, COL_NUM).value,
            "pedido": ws.cell(row, COL_PEDIDO).value,
            "ciudad": ws.cell(row, COL_CIUDAD).value,
            "deporte": ws.cell(row, COL_DEPORTE).value,
            "equipo": equipo,
            "rival": rival,
            "fecha": fecha,
            "hora": ws.cell(row, COL_HORA).value,
        })
    return matches


def load_known_contacts(wb) -> dict[str, dict]:
    """
    Lee la hoja EQUIPOS VISITANTES y devuelve un dict
    { nombre_normalizado: {telefono, email, contacto} }
    """
    ws = wb["EQUIPOS VISITANTES"]
    contacts = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name:
            continue
        telefono = ws.cell(row, 5).value
        email1 = ws.cell(row, 6).value
        email2 = ws.cell(row, 7).value
        email = email1 or email2
        contacto = ws.cell(row, 3).value

        # Normalizar nombre para búsqueda
        key = _normalize(str(name))
        contacts[key] = {
            "nombre": str(name),
            "telefono": str(telefono) if telefono else None,
            "email": str(email) if email else None,
            "contacto": str(contacto) if contacto else None,
        }
    return contacts


# Columnas en la hoja EQUIPOS VISITANTES (1-indexed)
CONT_COL_NOMBRE   = 1   # A: nombre del club
CONT_COL_CONTACTO = 3   # C: persona/departamento de contacto
CONT_COL_TEL      = 5   # E: teléfono
CONT_COL_EMAIL    = 6   # F: email principal
CONT_COL_WEB      = 8   # H: web / notas (si existe la columna)


def _find_last_contact_row(ws) -> int:
    last = 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, CONT_COL_NOMBRE).value is not None:
            last = row
    return last


def append_contact_row(wb, contact: dict) -> int:
    """
    Añade un contacto nuevo a la hoja EQUIPOS VISITANTES para cachearlo.
    contact keys: nombre, contacto, telefono, email, web
    Devuelve el número de fila escrito.
    """
    ws = wb["EQUIPOS VISITANTES"]
    next_row = _find_last_contact_row(ws) + 1

    ws.cell(next_row, CONT_COL_NOMBRE).value = contact.get("nombre", "")
    if contact.get("contacto"):
        ws.cell(next_row, CONT_COL_CONTACTO).value = contact["contacto"]
    if contact.get("telefono"):
        ws.cell(next_row, CONT_COL_TEL).value = contact["telefono"]
    if contact.get("email"):
        ws.cell(next_row, CONT_COL_EMAIL).value = contact["email"]
    # La web solo si la hoja tiene una columna preparada (evita machacar fórmulas)
    web = contact.get("web")
    if web and ws.max_column >= CONT_COL_WEB:
        cell = ws.cell(next_row, CONT_COL_WEB)
        if cell.value in (None, ""):
            cell.value = web

    return next_row


def match_already_registered(rival: str, fecha: datetime, equipo: str, existing: list[dict]) -> bool:
    """Devuelve True si el partido ya existe en el REGISTRO."""
    rival_norm = _normalize(rival)
    for m in existing:
        if _normalize(str(m.get("rival", ""))) != rival_norm:
            continue
        if m.get("equipo") and _normalize(str(m["equipo"])) != _normalize(equipo):
            continue
        m_fecha = m.get("fecha")
        if m_fecha and isinstance(m_fecha, datetime):
            if abs((m_fecha - fecha).days) <= 1:
                return True
    return False


def append_match_row(wb, match: dict) -> int:
    """
    Añade un partido nuevo en la siguiente fila libre del REGISTRO.
    Devuelve el número de fila donde se escribió.
    match keys: ciudad, deporte, equipo, rival, fecha (datetime), hora (float),
                lugar, telefono, email
    """
    ws = wb["REGISTRO"]
    next_row = _find_last_data_row(ws) + 1

    ws.cell(next_row, COL_PEDIDO).value = "PDTE."
    ws.cell(next_row, COL_CIUDAD).value = match.get("ciudad", "")
    ws.cell(next_row, COL_DEPORTE).value = match.get("deporte", "")
    ws.cell(next_row, COL_EQUIPO).value = match.get("equipo", "")
    ws.cell(next_row, COL_RIVAL).value = match.get("rival", "")

    fecha = match.get("fecha")
    if fecha:
        ws.cell(next_row, COL_FECHA).value = fecha.replace(tzinfo=None) if hasattr(fecha, "tzinfo") else fecha

    hora = match.get("hora")
    if hora:
        ws.cell(next_row, COL_HORA).value = hora

    if match.get("lugar"):
        ws.cell(next_row, COL_LUGAR).value = match["lugar"]
    if match.get("telefono"):
        ws.cell(next_row, COL_TEL).value = match["telefono"]
    if match.get("email"):
        ws.cell(next_row, COL_EMAIL).value = match["email"]

    return next_row


def save_workbook(wb, path: str):
    """Guarda el workbook. Crea backup .bak antes de sobreescribir."""
    bak = path + ".bak"
    try:
        shutil.copy2(path, bak)
    except Exception:
        pass
    wb.save(path)


def _normalize(s: str) -> str:
    """Normaliza nombre de equipo para comparación."""
    return s.upper().strip().replace("  ", " ")
